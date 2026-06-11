import os
import pandas as pd

from core.csv_loader import read_csv_safe
from core.schema import EXPECTED_COLUMNS
from core.data_cleaner import clean_value

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError


# ==================================================
# Operation Color Mapping
# ==================================================

COLOR_MAPPING = {

    # Unlock Success
    "UL.Success": PatternFill(
        start_color="FFA500",
        end_color="FFA500",
        fill_type="solid"
    ),

    "UL.Success.NFP": PatternFill(
        start_color="FFD27F",
        end_color="FFD27F",
        fill_type="solid"
    ),

    # Unlock Fail
    "UL.Fail": PatternFill(
        start_color="ADD8E6",
        end_color="ADD8E6",
        fill_type="solid"
    ),

    "UL.Fail(ensure_process)": PatternFill(
        start_color="B0E0E6",
        end_color="B0E0E6",
        fill_type="solid"
    ),

    # Token Success
    "T.Success": PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    ),

    # Fraud Related
    "AL.Unlocked": PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    ),

    "P.Fail": PatternFill(
        start_color="FF6666",
        end_color="FF6666",
        fill_type="solid"
    ),

    "F.Reboot": PatternFill(
        start_color="D8B4FE",
        end_color="D8B4FE",
        fill_type="solid"
    ),

    "UL.Airwatt short": PatternFill(
        start_color="F9A8D4",
        end_color="F9A8D4",
        fill_type="solid"
    )
}


# ==================================================
# CSV MERGER
# ==================================================

def merge_customer_data(customer_results):

    all_data = []

    stats = {
        "merged_files": 0,
        "skipped_unlock": 0,
        "failed_files": 0,
        "errors": []
    }

    for customer in customer_results:

        ub3 = customer["ub3_folder"]
        customer_id = customer["customer_id"]

        for file_path in customer["csv_files"]:

            filename = os.path.basename(
                file_path
            ).upper()

            # Skip Unlock Files

            if filename.startswith(
                "UNLOCK"
            ):

                stats["skipped_unlock"] += 1
                continue

            try:

                df = read_csv_safe(
                    file_path
                )

                # Ensure schema consistency

                for col in EXPECTED_COLUMNS:

                    if col not in df.columns:
                        df[col] = None

                df = df[
                    EXPECTED_COLUMNS
                ]

                # Metadata

                df["UB3"] = ub3

                df["CustomerID"] = (
                    customer_id
                )

                df["SourceFile"] = (
                    os.path.basename(
                        file_path
                    )
                )

                all_data.append(
                    df
                )

                stats["merged_files"] += 1

            except Exception as e:

                stats["failed_files"] += 1

                stats["errors"].append(
                    f"{os.path.basename(file_path)} : {str(e)}"
                )

    if not all_data:

        return None, stats

    merged_df = pd.concat(
        all_data,
        ignore_index=True
    )

    return merged_df, stats


# ==================================================
# EXCEL EXPORT
# ==================================================

def export_excel(df, output_path):

    wb = Workbook()

    ws = wb.active

    ws.title = "Merged Data"

    # =====================================
    # Header
    # =====================================

    ws.append(
        list(df.columns)
    )

    operation_col = None

    if "Operation" in df.columns:

        operation_col = (
            list(df.columns)
            .index("Operation")
            + 1
        )

    # =====================================
    # Data Rows
    # =====================================

    for row_idx, row in enumerate(
        df.itertuples(index=False),
        start=2
    ):

        cleaned_row = []

        for col_idx, value in enumerate(row):

            try:

                cleaned_value = clean_value(
                    value
                )

                cleaned_row.append(
                    cleaned_value
                )

            except Exception:

                print(
                    f"Clean Error -> "
                    f"Row:{row_idx} "
                    f"Col:{df.columns[col_idx]} "
                    f"Value:{repr(value)}"
                )

                raise

        try:

            ws.append(
                cleaned_row
            )

        except IllegalCharacterError:

            print(
                "\n=== BAD DATA FOUND ==="
            )

            print(
                f"Row Number: {row_idx}"
            )

            for col_name, value in zip(
                df.columns,
                cleaned_row
            ):

                print(
                    f"{col_name}: {repr(value)}"
                )

            raise

        # =====================================
        # Row Color Coding
        # =====================================

        if operation_col:

            operation_value = str(
                cleaned_row[
                    operation_col - 1
                ]
            ).strip()

            fill = COLOR_MAPPING.get(
                operation_value
            )

            if fill:

                for col in range(
                    1,
                    len(cleaned_row) + 1
                ):

                    ws.cell(
                        row=row_idx,
                        column=col
                    ).fill = fill

    # =====================================
    # Auto Width
    # =====================================

    for column in ws.columns:

        max_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            try:

                length = len(
                    str(cell.value)
                )

                if length > max_length:
                    max_length = length

            except:
                pass

        adjusted_width = min(
            max_length + 3,
            60
        )

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    # =====================================
    # Save Workbook
    # =====================================

    wb.save(
        output_path
    )